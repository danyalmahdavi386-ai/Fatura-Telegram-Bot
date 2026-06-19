from __future__ import annotations

import asyncio
import logging
import os
import re
import tempfile
import threading
import uuid
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from PIL import Image, ImageOps
from pydantic import BaseModel, Field
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.constants import ChatAction
from telegram.ext import (
    Application,
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

IS_RAILWAY = bool(
    os.getenv("RAILWAY_ENVIRONMENT")
    or os.getenv("RAILWAY_PROJECT_ID")
    or os.getenv("RAILWAY_SERVICE_ID")
)
DEFAULT_DATA_DIR = "/data" if IS_RAILWAY else str(BASE_DIR / "data")

DATA_DIR = Path(os.getenv("DATA_DIR", DEFAULT_DATA_DIR)).expanduser()
LOG_DIR = Path(os.getenv("LOG_DIR", str(DATA_DIR / "logs"))).expanduser()
MONTHLY_DIR = DATA_DIR / "monthly"
LEGACY_EXCEL_PATH = DATA_DIR / "invoices.xlsx"
ISTANBUL_TZ = ZoneInfo("Europe/Istanbul")

DATA_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)
MONTHLY_DIR.mkdir(parents=True, exist_ok=True)

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash").strip()
ALLOWED_USER_ID_RAW = os.getenv("ALLOWED_USER_ID", "").strip()
ALLOWED_USER_ID = int(ALLOWED_USER_ID_RAW) if ALLOWED_USER_ID_RAW else None

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "bot.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("invoice_bot")

excel_lock = threading.Lock()
gemini_client = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None


class InvoiceItem(BaseModel):
    description: Optional[str] = Field(
        default=None, description="Exact item or service description as printed."
    )
    quantity: Optional[float] = Field(default=None, description="Item quantity.")
    unit: Optional[str] = Field(default=None, description="Unit such as adet, kg, m, saat.")
    unit_price: Optional[float] = Field(default=None, description="Unit price before VAT if clear.")
    vat_rate: Optional[float] = Field(default=None, description="VAT rate as a percentage, e.g. 20.")
    line_total: Optional[float] = Field(default=None, description="Line total as printed.")


class InvoiceData(BaseModel):
    document_type: Optional[str] = Field(
        default=None, description="Invoice, receipt, expense slip, or other."
    )
    supplier_name: Optional[str] = None
    supplier_tax_number: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = Field(
        default=None, description="Keep the date exactly as printed."
    )
    invoice_date_iso: Optional[str] = Field(
        default=None,
        description="Normalized invoice date in YYYY-MM-DD format when visible.",
    )
    currency: Optional[str] = Field(
        default=None, description="TRY, USD, EUR, GBP, or another visible currency."
    )
    subtotal: Optional[float] = None
    vat_total: Optional[float] = None
    grand_total: Optional[float] = None
    items: list[InvoiceItem] = Field(default_factory=list)
    confidence: float = Field(
        ge=0, le=100, description="Overall extraction confidence from 0 to 100."
    )
    notes: Optional[str] = Field(
        default=None,
        description="Unreadable, ambiguous, inconsistent, or suspicious fields. Do not invent data.",
    )


EXTRACTION_PROMPT = """
You are a careful invoice data extraction system for personal bookkeeping.

The document may be Turkish or multilingual. Extract only information visibly supported
by the image. Never guess missing numbers. Use null for unreadable or absent values.

Rules:
1. Preserve invoice number, tax number, supplier name, and invoice_date exactly as printed.
2. Also return invoice_date_iso in YYYY-MM-DD format when the date is readable.
3. Monetary values must be numbers without currency symbols or thousands separators.
4. Identify subtotal, VAT total, and grand total separately.
5. Extract line items when they are readable.
6. VAT rate is a percentage such as 1, 10, or 20.
7. If totals do not reconcile, mention this in notes.
8. If the image is not an invoice/receipt, explain it in notes and keep unknown fields null.
9. Return a realistic overall confidence score.
"""


def authorized_user(user_id: Optional[int]) -> bool:
    return ALLOWED_USER_ID is not None and user_id == ALLOWED_USER_ID


def clean_text(value: object) -> str:
    if value is None:
        return "—"
    text = str(value).strip()
    return text if text else "—"


def money_text(value: Optional[float], currency: Optional[str]) -> str:
    if value is None:
        return "—"
    suffix = f" {currency}" if currency else ""
    return f"{value:,.2f}{suffix}"


def prepare_image(input_path: Path, output_path: Path) -> None:
    """Correct rotation, resize large images, and save as a clean JPEG."""
    with Image.open(input_path) as image:
        image = ImageOps.exif_transpose(image)
        if image.mode != "RGB":
            image = image.convert("RGB")
        max_side = 2200
        if max(image.size) > max_side:
            image.thumbnail((max_side, max_side))
        image.save(output_path, format="JPEG", quality=92, optimize=True)



def extract_invoice(image_path: Path) -> InvoiceData:
    if gemini_client is None:
        raise RuntimeError("GEMINI_API_KEY تنظیم نشده است.")

    image_bytes = image_path.read_bytes()
    image_part = types.Part.from_bytes(
        data=image_bytes,
        mime_type="image/jpeg",
    )

    response = gemini_client.models.generate_content(
        model=GEMINI_MODEL,
        contents=[
            image_part,
            EXTRACTION_PROMPT,
            "Extract this invoice or receipt into the required JSON structure.",
        ],
        config={
            "temperature": 0.1,
            "response_format": {
                "text": {
                    "mime_type": "application/json",
                    "schema": InvoiceData.model_json_schema(),
                }
            },
        },
    )

    if not response.text:
        raise RuntimeError("جمنای خروجی قابل‌خواندن تولید نکرد.")

    try:
        return InvoiceData.model_validate_json(response.text)
    except Exception as exc:
        logger.error("Invalid Gemini JSON: %s", response.text)
        raise RuntimeError("خروجی جمنای با ساختار فاکتور سازگار نبود.") from exc


TURKISH_MONTHS = {
    "ocak": 1,
    "şubat": 2,
    "subat": 2,
    "mart": 3,
    "nisan": 4,
    "mayıs": 5,
    "mayis": 5,
    "haziran": 6,
    "temmuz": 7,
    "ağustos": 8,
    "agustos": 8,
    "eylül": 9,
    "eylul": 9,
    "ekim": 10,
    "kasım": 11,
    "kasim": 11,
    "aralık": 12,
    "aralik": 12,
}


def now_istanbul() -> datetime:
    return datetime.now(ISTANBUL_TZ)


def parse_month_key(invoice: InvoiceData) -> tuple[str, bool]:
    """
    Return YYYY-MM and whether processing date was used as a fallback.
    Priority:
    1) Gemini normalized invoice_date_iso
    2) Common numeric/text formats in invoice_date
    3) Processing month in Europe/Istanbul
    """
    if invoice.invoice_date_iso:
        match = re.fullmatch(
            r"\s*(\d{4})-(\d{1,2})-(\d{1,2})\s*",
            invoice.invoice_date_iso,
        )
        if match:
            year, month, _ = map(int, match.groups())
            if 1 <= month <= 12:
                return f"{year:04d}-{month:02d}", False

    raw_date = (invoice.invoice_date or "").strip().lower()

    # YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD
    match = re.search(
        r"\b(\d{4})[./-](\d{1,2})[./-](\d{1,2})\b",
        raw_date,
    )
    if match:
        year, month, _ = map(int, match.groups())
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}", False

    # DD-MM-YYYY, DD/MM/YYYY, DD.MM.YYYY
    match = re.search(
        r"\b(\d{1,2})[./-](\d{1,2})[./-](\d{4})\b",
        raw_date,
    )
    if match:
        _, month, year = map(int, match.groups())
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}", False

    # Turkish dates such as: 19 Haziran 2026
    match = re.search(
        r"\b\d{1,2}\s+([a-zçğıöşü]+)\s+(\d{4})\b",
        raw_date,
        re.IGNORECASE,
    )
    if match:
        month_name, year_text = match.groups()
        month = TURKISH_MONTHS.get(month_name.lower())
        if month:
            return f"{int(year_text):04d}-{month:02d}", False

    current = now_istanbul()
    return current.strftime("%Y-%m"), True


def monthly_excel_path(month_key: str) -> Path:
    return MONTHLY_DIR / f"invoices_{month_key}.xlsx"


def ensure_workbook(excel_path: Path) -> None:
    if excel_path.exists():
        return

    workbook = Workbook()
    invoices_sheet = workbook.active
    invoices_sheet.title = "فاکتورها"
    invoices_sheet.sheet_view.rightToLeft = True
    invoices_sheet.freeze_panes = "A2"

    invoice_headers = [
        "شناسه ثبت",
        "زمان پردازش",
        "نوع سند",
        "نام فروشنده",
        "شماره مالیاتی",
        "شماره فاکتور",
        "تاریخ فاکتور",
        "تاریخ استاندارد",
        "ارز",
        "مبلغ بدون مالیات",
        "مالیات",
        "مبلغ نهایی",
        "درصد اطمینان",
        "یادداشت",
        "نام فایل",
        "شناسه کاربر",
    ]
    invoices_sheet.append(invoice_headers)

    items_sheet = workbook.create_sheet("اقلام")
    items_sheet.sheet_view.rightToLeft = True
    items_sheet.freeze_panes = "A2"
    item_headers = [
        "شناسه ثبت",
        "شماره ردیف",
        "شرح",
        "تعداد",
        "واحد",
        "قیمت واحد",
        "نرخ مالیات",
        "مبلغ ردیف",
    ]
    items_sheet.append(item_headers)

    for sheet in (invoices_sheet, items_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.auto_filter.ref = sheet.dimensions

    workbook.save(excel_path)


def autosize_worksheet(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 10),
            45,
        )


def workbook_stats(excel_path: Path) -> tuple[int, int, int]:
    with excel_lock:
        ensure_workbook(excel_path)
        workbook = load_workbook(
            excel_path,
            read_only=True,
            data_only=True,
        )
        invoice_rows = max(workbook["فاکتورها"].max_row - 1, 0)
        item_rows = max(workbook["اقلام"].max_row - 1, 0)
        workbook.close()
        file_size = excel_path.stat().st_size
        return invoice_rows, item_rows, file_size


def append_invoice_to_excel(
    invoice: InvoiceData,
    source_name: str,
    user_id: int,
) -> tuple[str, Path, str, int, int, bool]:
    month_key, used_fallback = parse_month_key(invoice)
    excel_path = monthly_excel_path(month_key)

    with excel_lock:
        ensure_workbook(excel_path)
        workbook = load_workbook(excel_path)
        invoices_sheet = workbook["فاکتورها"]
        items_sheet = workbook["اقلام"]

        record_id = uuid.uuid4().hex[:12].upper()
        processed_at = now_istanbul().strftime("%Y-%m-%d %H:%M:%S")

        invoices_sheet.append(
            [
                record_id,
                processed_at,
                invoice.document_type,
                invoice.supplier_name,
                invoice.supplier_tax_number,
                invoice.invoice_number,
                invoice.invoice_date,
                invoice.invoice_date_iso,
                invoice.currency,
                invoice.subtotal,
                invoice.vat_total,
                invoice.grand_total,
                invoice.confidence,
                invoice.notes,
                source_name,
                user_id,
            ]
        )

        for index, item in enumerate(invoice.items, start=1):
            items_sheet.append(
                [
                    record_id,
                    index,
                    item.description,
                    item.quantity,
                    item.unit,
                    item.unit_price,
                    item.vat_rate,
                    item.line_total,
                ]
            )

        for sheet in (invoices_sheet, items_sheet):
            sheet.auto_filter.ref = sheet.dimensions
            autosize_worksheet(sheet)

        for row in invoices_sheet.iter_rows(min_row=2, min_col=10, max_col=12):
            for cell in row:
                cell.number_format = "#,##0.00"

        for row in items_sheet.iter_rows(min_row=2, min_col=4, max_col=8):
            for cell in row:
                if cell.column != 5:
                    cell.number_format = "#,##0.00"

        temp_path = excel_path.with_suffix(".tmp.xlsx")
        workbook.save(temp_path)
        os.replace(temp_path, excel_path)

        invoice_count = max(invoices_sheet.max_row - 1, 0)
        item_count = max(items_sheet.max_row - 1, 0)

    return (
        record_id,
        excel_path,
        month_key,
        invoice_count,
        item_count,
        used_fallback,
    )


def available_months() -> list[str]:
    result = []
    for path in MONTHLY_DIR.glob("invoices_????-??.xlsx"):
        match = re.fullmatch(r"invoices_(\d{4}-\d{2})\.xlsx", path.name)
        if match:
            result.append(match.group(1))
    return sorted(set(result), reverse=True)


def normalize_requested_month(value: str) -> Optional[str]:
    value = value.strip()
    match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", value)
    if not match:
        return None
    year, month = map(int, match.groups())
    if not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}"

def invoice_preview(invoice: InvoiceData) -> str:
    item_count = len(invoice.items)
    notes = clean_text(invoice.notes)
    if len(notes) > 500:
        notes = notes[:500] + "…"

    return (
        "✅ اطلاعات استخراج شد\n\n"
        f"نوع سند: {clean_text(invoice.document_type)}\n"
        f"فروشنده: {clean_text(invoice.supplier_name)}\n"
        f"شماره مالیاتی: {clean_text(invoice.supplier_tax_number)}\n"
        f"شماره فاکتور: {clean_text(invoice.invoice_number)}\n"
        f"تاریخ: {clean_text(invoice.invoice_date)}\n"
        f"تاریخ استاندارد: {clean_text(invoice.invoice_date_iso)}\n"
        f"مبلغ بدون مالیات: {money_text(invoice.subtotal, invoice.currency)}\n"
        f"مالیات: {money_text(invoice.vat_total, invoice.currency)}\n"
        f"مبلغ نهایی: {money_text(invoice.grand_total, invoice.currency)}\n"
        f"تعداد اقلام خوانده‌شده: {item_count}\n"
        f"اطمینان: {invoice.confidence:.0f}٪\n"
        f"یادداشت: {notes}\n\n"
        "برای ثبت در اکسل تأیید کن."
    )


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id if update.effective_user else None

    if ALLOWED_USER_ID is None:
        message = (
            "سلام. بات هنوز برای کاربر مجاز تنظیم نشده است.\n\n"
            "۱) دستور /myid را بفرست.\n"
            "۲) عدد نمایش‌داده‌شده را در فایل .env مقابل ALLOWED_USER_ID قرار بده.\n"
            "۳) بات را دوباره اجرا کن."
        )
    elif authorized_user(user_id):
        message = (
            "سلام 👋\n"
            "یک عکس واضح از فاکتور یا رسید بفرست.\n"
            "من اطلاعات را استخراج می‌کنم، قبل از ثبت پیش‌نمایش می‌دهم "
            "و پس از تأیید آن را به فایل همان ماه اضافه می‌کنم.\n\n"
            "دستورها:\n"
            "/export دریافت فایل کامل ماه جاری\n"
            "/export 2026-06 دریافت فایل کامل یک ماه مشخص\n"
            "/months نمایش ماه‌های موجود\n"
            "/status وضعیت فایل ماه جاری\n"
            "/cancel لغو فاکتور در انتظار\n"
            "/myid نمایش شناسه تلگرام"
        )
    else:
        message = "این بات شخصی است و اجازه استفاده از آن را نداری."

    await update.effective_message.reply_text(message)


async def myid_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id if update.effective_user else None
    await update.effective_message.reply_text(f"شناسه تلگرام تو:\n{user_id}")


async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not authorized_user(update.effective_user.id if update.effective_user else None):
        await update.effective_message.reply_text("اجازه دسترسی نداری.")
        return
    context.user_data.pop("pending_invoice", None)
    context.user_data.pop("pending_source_name", None)
    await update.effective_message.reply_text("فاکتور در انتظار لغو شد.")


async def export_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    if not authorized_user(update.effective_user.id if update.effective_user else None):
        await update.effective_message.reply_text("اجازه دسترسی نداری.")
        return

    if context.args:
        month_key = normalize_requested_month(context.args[0])
        if month_key is None:
            await update.effective_message.reply_text(
                "فرمت ماه درست نیست.\n"
                "نمونه صحیح:\n/export 2026-06"
            )
            return
    else:
        month_key = now_istanbul().strftime("%Y-%m")

    excel_path = monthly_excel_path(month_key)
    if not excel_path.exists():
        await update.effective_message.reply_text(
            f"برای ماه {month_key} هنوز فایل یا فاکتوری ثبت نشده است."
        )
        return

    invoice_count, item_count, _ = await asyncio.to_thread(
        workbook_stats,
        excel_path,
    )

    with excel_path.open("rb") as file_handle:
        await update.effective_message.reply_document(
            document=file_handle,
            filename=excel_path.name,
            caption=(
                f"فایل فاکتورهای ماه {month_key}\n"
                f"تعداد فاکتورها: {invoice_count}\n"
                f"تعداد اقلام: {item_count}"
            ),
        )


async def months_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    if not authorized_user(update.effective_user.id if update.effective_user else None):
        await update.effective_message.reply_text("اجازه دسترسی نداری.")
        return

    months = available_months()
    if not months:
        await update.effective_message.reply_text(
            "هنوز هیچ فایل ماهانه‌ای ساخته نشده است."
        )
        return

    shown = months[:24]
    lines = [f"• {month}" for month in shown]
    message = (
        "📁 ماه‌های موجود:\n\n"
        + "\n".join(lines)
        + "\n\nبرای دریافت هر ماه بنویس:\n/export 2026-06"
    )
    await update.effective_message.reply_text(message)


async def status_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    if not authorized_user(update.effective_user.id if update.effective_user else None):
        await update.effective_message.reply_text("اجازه دسترسی نداری.")
        return

    month_key = now_istanbul().strftime("%Y-%m")
    excel_path = monthly_excel_path(month_key)

    if not excel_path.exists():
        await update.effective_message.reply_text(
            f"ماه جاری: {month_key}\n"
            "هنوز فاکتوری برای این ماه ثبت نشده است.\n"
            f"مسیر ذخیره: {excel_path}"
        )
        return

    invoice_count, item_count, file_size = await asyncio.to_thread(
        workbook_stats,
        excel_path,
    )

    await update.effective_message.reply_text(
        "📊 وضعیت فایل ماه جاری\n\n"
        f"ماه: {month_key}\n"
        f"تعداد فاکتورها: {invoice_count}\n"
        f"تعداد اقلام: {item_count}\n"
        f"حجم فایل: {file_size:,} بایت\n"
        f"مسیر ذخیره: {excel_path}\n"
        f"Railway: {'بله' if IS_RAILWAY else 'خیر'}"
    )


async def handle_invoice_file(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    message = update.effective_message
    user = update.effective_user

    if message is None or user is None:
        return

    if not authorized_user(user.id):
        await message.reply_text(
            "این بات هنوز برای شناسه تو فعال نشده است. ابتدا /myid را بفرست."
        )
        return

    if not GEMINI_API_KEY:
        await message.reply_text("کلید جمنای در متغیر GEMINI_API_KEY تنظیم نشده است.")
        return

    source_name = "telegram_photo.jpg"
    suffix = ".jpg"

    if message.photo:
        telegram_file = await message.photo[-1].get_file()
    elif message.document:
        mime_type = (message.document.mime_type or "").lower()
        allowed_types = {"image/jpeg", "image/png", "image/webp"}
        if mime_type not in allowed_types:
            await message.reply_text(
                "در نسخه اول فقط عکس‌های JPG، PNG و WEBP پشتیبانی می‌شوند."
            )
            return
        source_name = message.document.file_name or "invoice_image"
        suffix = Path(source_name).suffix or ".jpg"
        telegram_file = await message.document.get_file()
    else:
        return

    await message.chat.send_action(ChatAction.TYPING)
    status_message = await message.reply_text("در حال خواندن فاکتور…")

    try:
        with tempfile.TemporaryDirectory(prefix="invoice_bot_") as temp_dir:
            raw_path = Path(temp_dir) / f"raw{suffix}"
            prepared_path = Path(temp_dir) / "prepared.jpg"

            await telegram_file.download_to_drive(raw_path)
            await asyncio.to_thread(prepare_image, raw_path, prepared_path)
            invoice = await asyncio.to_thread(extract_invoice, prepared_path)

        context.user_data["pending_invoice"] = invoice.model_dump()
        context.user_data["pending_source_name"] = source_name

        keyboard = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("✅ ثبت و دریافت اکسل", callback_data="save_invoice"),
                    InlineKeyboardButton("❌ لغو", callback_data="cancel_invoice"),
                ]
            ]
        )
        await status_message.edit_text(invoice_preview(invoice), reply_markup=keyboard)

    except Exception as exc:
        logger.exception("Invoice processing failed: %s", exc)
        await status_message.edit_text(
            "پردازش فاکتور ناموفق بود.\n"
            "عکس واضح‌تر و مستقیم‌تری بفرست یا گزارش خطا را در logs/bot.log بررسی کن."
        )


async def invoice_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    query = update.callback_query
    user = update.effective_user

    if query is None or user is None:
        return

    await query.answer()

    if not authorized_user(user.id):
        await query.edit_message_text("اجازه دسترسی نداری.")
        return

    if query.data == "cancel_invoice":
        context.user_data.pop("pending_invoice", None)
        context.user_data.pop("pending_source_name", None)
        await query.edit_message_text("فاکتور ثبت نشد.")
        return

    if query.data != "save_invoice":
        return

    pending = context.user_data.get("pending_invoice")
    source_name = context.user_data.get("pending_source_name", "telegram_photo.jpg")

    if not pending:
        await query.edit_message_text(
            "اطلاعات فاکتور در حافظه نیست. لطفاً عکس را دوباره بفرست."
        )
        return

    try:
        invoice = InvoiceData.model_validate(pending)
        (
            record_id,
            excel_path,
            month_key,
            invoice_count,
            item_count,
            used_fallback,
        ) = await asyncio.to_thread(
            append_invoice_to_excel,
            invoice,
            source_name,
            user.id,
        )

        context.user_data.pop("pending_invoice", None)
        context.user_data.pop("pending_source_name", None)

        fallback_text = (
            "\n⚠️ تاریخ فاکتور قابل تشخیص نبود؛ ماه پردازش استفاده شد."
            if used_fallback
            else ""
        )

        await query.edit_message_text(
            "✅ فاکتور در فایل اصلی همان ماه ثبت شد.\n"
            f"ماه فایل: {month_key}\n"
            f"شناسه ثبت: {record_id}\n"
            f"تعداد کل فاکتورهای این ماه: {invoice_count}\n"
            f"تعداد کل اقلام این ماه: {item_count}\n\n"
            f"برای دریافت فایل کامل بنویس:\n/export {month_key}"
            f"{fallback_text}"
        )
    except Exception as exc:
        logger.exception("Excel save failed: %s", exc)
        await query.edit_message_text(
            "ثبت در اکسل ناموفق بود. گزارش خطا در logs/bot.log ذخیره شد."
        )


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.exception("Unhandled Telegram error", exc_info=context.error)


def validate_settings() -> None:
    missing = []
    if not TELEGRAM_BOT_TOKEN:
        missing.append("TELEGRAM_BOT_TOKEN")
    if not GEMINI_API_KEY:
        missing.append("GEMINI_API_KEY")
    if missing:
        raise RuntimeError(
            "این متغیرها در فایل .env تنظیم نشده‌اند: " + ", ".join(missing)
        )


def main() -> None:
    validate_settings()

    application: Application = (
        ApplicationBuilder()
        .token(TELEGRAM_BOT_TOKEN)
        .build()
    )

    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("myid", myid_command))
    application.add_handler(CommandHandler("status", status_command))
    application.add_handler(CommandHandler("months", months_command))
    application.add_handler(CommandHandler("export", export_command))
    application.add_handler(CommandHandler("cancel", cancel_command))
    application.add_handler(CallbackQueryHandler(invoice_callback))
    application.add_handler(
        MessageHandler(filters.PHOTO | filters.Document.ALL, handle_invoice_file)
    )
    application.add_error_handler(error_handler)

    logger.info(
        "Bot started | model=%s | railway=%s | monthly_dir=%s",
        GEMINI_MODEL,
        IS_RAILWAY,
        MONTHLY_DIR,
    )
    application.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
